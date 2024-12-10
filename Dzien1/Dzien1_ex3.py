import openpyxl

wb = openpyxl.load_workbook('../data/videogamesales.xlsx')
ws = wb.active

ws = wb['vgsales']

row_pos = 2
col_pos = 7

total_sales = ((ws.cell(row=row_pos, column=col_pos).value)
               + (ws.cell(row=row_pos, column=col_pos + 1).value)
               + (ws.cell(row=row_pos, column=col_pos + 2).value)
               + (ws.cell(row=row_pos, column=col_pos + 3).value))

ws.cell(row=2,column=11).value=total_sales
wb.save('videogamesales.xlsx')

