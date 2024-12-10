import openpyxl

wb = openpyxl.load_workbook('../data/videogamesales.xlsx')
# wb = openpyxl.load_workbook('videogamesales.xlsx') jakby plik byl zapisany w tym samym folderze co kod python

ws = wb.active

ws = wb['Video Game Sales Data']

# dodawanie kolumny z total_sales
row_pos=1
for i in range(1, ws.max_row):
    row_pos +=1
    NA_Sales = ws.cell(row=row_pos, column=7).value
    EU_Sales = ws.cell(row=row_pos, column=8).value
    JP_Sales = ws.cell(row=row_pos, column=9).value
    Other_Sales = ws.cell(row=row_pos, column=10).value

    total_sales = NA_Sales + EU_Sales + JP_Sales + Other_Sales
    ws.cell(row=row_pos, column=11).value = total_sales


# new_row=(1,"The legend",1984,"Action","Nintendo",120,80,6.76,2.5,20,2.5) # dodawanie wiersza
# ws.append(new_row)

#wb.save('../data/videogamesales.xlsx') # ten plik zapisal sie w folderze Dzien1

values=[ws.cell(row=ws.max_row,column=i).value for i in range(1,ws.max_column+1)]
print(values)

#ws.delete_rows(ws.max_row,amount=) # usuwanie ostatniego wiersza
#ws.delete_rows(ws.max_row-1,amount=2) # usunal 2 ostatnie wiersze

ws['P1']="Average value"
ws['P2']="=average(K2:K16220)"

a=str(ws.max_row)
print(a)
ws['Q1']="Number of calls"
ws['Q2']="=counta(E2:E" + a + ")"

ws['R1']="Number of cells with Sports Game"
ws['R2']='=countif(E2:E16220,"Sports")'

ws['S1']="Total sports Sales"
ws['S2']='=sumifs(K2:K16220,E2:E16220,"Sports")'

# ws['T1']="Total rows"
# ws['T2']='=counta(A:A)'

ws['T1']="Rounded sum of Sports Sales"
ws['T2']='=ceiling(S2,25)'

# print(ws.title)
# ws.title="Video Game Sales Data"

print(wb.sheetnames) # wszystkie nazwy arkuszy

# wb.create_sheet('Empty Sheet') # tworzenie nowego arkusza
# print(wb.sheetnames) # sprawdzenie czy arkusz został utworzony

wb.remove(wb['Empty Sheet1'])

wb.copy_worksheet(wb['Video Game Sales Data'])
print(wb.sheetnames) # wszystkie nazwy arkuszy

wb.save('../data/videogamesales.xlsx')



