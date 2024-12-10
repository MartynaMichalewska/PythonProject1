

import openpyxl
from openpyxl.chart import Reference,BarChart
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('../data/videogamesales.xlsx')
ws=wb.active

ws=wb['Video Game Sales Data']

# ws['A1'].font=Font(bold=True, size=12) # zmiana formatowania dla komorki A1
ws['A1'].font=Font(color='FF0000',bold=True, size=12) # zmiana koloru dla komorki A1, czerwony
ws['A2'].font=Font(color='0000FF',bold=True, size=12) # zmiana koloru dla komorki A1, niebieski

ws['A1'].fill=PatternFill(patternType='solid', start_color="38e3ff") # wypełnienie komórki A1

my_border=Side(border_style='thin', color='000000')
ws['A1'].border=Border(top=my_border, left=my_border, right=my_border, bottom=my_border) # wstawianie obramowania dla komórki

# for cell in ws['1:1']:
#     cell.font=Font(bold=True, size=12) # zmiana formatowania dla wszystkich komórek w wierszu 1, ale dziala tylko n atych ktore byly uzywane

fill=PatternFill(patternType='solid', start_color="90EE90", end_color="90EE90") # wypełnienie całego arkusza
ws.conditional_formatting.add('G2:K16594',CellIsRule(operator='greaterThan', formula=[8], fill=fill, font=Font(color="FF0000"))) #  wstawianie warunkowego wypełnienia, maluje wieksze od 8



wb.save('../data/videogamesales.xlsx')
