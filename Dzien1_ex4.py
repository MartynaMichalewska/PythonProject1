import openpyxl

wb = openpyxl.load_workbook('videogamesales.xlsx')
ws = wb.active

ws = wb['vgsales']

row_pos = 1
for i in range(1, ws.max_row):
    row_pos +=1
    NA_Sales = ws.cell(row=row_pos, column=7).value
    EU_Sales = ws.cell(row=row_pos, column=8).value
    JP_Sales = ws.cell(row=row_pos, column=9).value
    Other_Sales = ws.cell(row=row_pos, column=10).value

    total_sales = NA_Sales + EU_Sales + JP_Sales + Other_Sales
    ws.cell(row=row_pos, column=11).value = total_sales

wb.save('videogamesales.xlsx')