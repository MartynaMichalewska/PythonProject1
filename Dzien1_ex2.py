import openpyxl

wb=openpyxl.load_workbook('videogamesales.xlsx')
ws=wb.active

ws=wb['vgsales']

print(f'Total number of rows + {ws.max_row}')
print('Total number of rows ' + str(ws.max_row))

values=[ws.cell(row=1,column=i).value for i in range(1,ws.max_column)]
print(values)

data=[ws.cell(row=i,column=2).value for i in range(2,12)]
print(data)

my_list=list()

for  value in ws.iter_rows(
        min_row=1, max_row=11, min_col=1, max_col=6,
        values_only=True):
    my_list.append(value)

for ele1, ele2, ele3, ele4, ele5, ele6 in my_list:
    print("{:<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1, ele2, ele3, ele4, ele5, ele6))
