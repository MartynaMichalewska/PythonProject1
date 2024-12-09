# import openpyxl
from openpyxl import Workbook, load_workbook
wb=Workbook()
ws=wb.active

ws['A1']=42
ws.append([1,2,3])
wb.save('sample.xlsx')

# wprowadz dane do komowrki w juz istniejacym pliku

# wb1=load_workbook("C:\Users\cscomarch\Desktop\Szkolenie\przyklad PQ\PQoutput")

workbook=load_workbook('sample.xlsx')
sheet=workbook.active
print(sheet)
print(sheet['A1'].value)

for row in sheet.iter_rows(min_row=1, max_row=2):
    for cell in row:
        print(cell.value)


# workbook1=load_workbook(r'C:\Users\cscomarch\Desktop\Szkolenie\przyklad PQ\Zeszyt1 — kopia.xlsx')
# # sheet1=workbook.active
# sheet1=workbook1.active
# sheet1=workbook1['Arkusz1']
# sheet1['D2']="testowy komentarz"
# # sheet1.append(['testowy komentarz'])
# workbook1.save(r'C:\Users\cscomarch\Desktop\Szkolenie\przyklad PQ\Zeszyt1 — kopia.xlsx')

workbook1=load_workbook(r'C:\Users\cscomarch\Desktop\Szkolenie\przyklad PQ\PQoutput_backup — kopia.xlsx')
# sheet1=workbook.active
sheet1=workbook1.active
sheet1=workbook1['Arkusz2']
sheet1['G2']="testowy komentarz"
# sheet1.append(['testowy komentarz'])
workbook1.save(r'C:\Users\cscomarch\Desktop\Szkolenie\przyklad PQ\PQoutput_backup — kopia.xlsx')







